#!/usr/bin/env python3
import os
import sys

# Adjust path to import src modules
ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(ROOT_DIR)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl no está instalado. Instálalo con 'pip install openpyxl'")
    sys.exit(1)

def build_excel_checklist():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styling definitions
    green_header_fill = PatternFill(start_color="00594E", end_color="00594E", fill_type="solid")
    gold_fill = PatternFill(start_color="B5A160", end_color="B5A160", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F4", end_color="F2F5F4", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFDF6", end_color="FFFDF6", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    font_title = Font(name="Calibri", size=16, bold=True, color="00594E")
    font_section_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_italic = Font(name="Calibri", size=10, italic=True)

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    thick_bottom_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='medium', color='00594E')
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ─────────────────────────────────────────────────────────────────────────
    # HOJA 1: CREDENCIALES
    # ─────────────────────────────────────────────────────────────────────────
    ws_cred = wb.create_sheet(title="1. Credenciales de Acceso")
    ws_cred.views.sheetView[0].showGridLines = True

    # Title block
    ws_cred["A1"] = "🔑 Credenciales de Acceso - Plataforma PA_TH"
    ws_cred["A1"].font = font_title
    ws_cred.row_dimensions[1].height = 30

    ws_cred["A2"] = "Usa estas cuentas de prueba para iniciar sesión y verificar el menú según cada rol:"
    ws_cred["A2"].font = font_italic
    ws_cred.row_dimensions[2].height = 20

    # Table Headers
    headers_cred = ["Nombre / Cargo", "Rol en Sistema", "Usuario Streamlit", "Contraseña de Prueba"]
    for col_idx, h in enumerate(headers_cred, 1):
        cell = ws_cred.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = green_header_fill
        cell.alignment = align_center
        cell.border = thick_bottom_border
    ws_cred.row_dimensions[4].height = 28

    # Data
    users_data = [
        ("Administrador General", "Admin", "admin", "admin123"),
        ("Lilia Andrea Nocua Neme (Jefe de Oficina)", "Supervisor", "lilia.nocua", "lilia123"),
        ("Rosaura Andrea Ramirez (Profesional TH)", "Worker", "rosaura.ramirez", "rosaura123"),
        ("Laura Vanesa Cely (Profesional TH)", "Worker", "laura.cely", "laura123"),
        ("Claudia America Garcia (Profesional TH)", "Worker", "claudia.garcia", "claudia123"),
        ("Leidy Yamile Garcia (Profesional TH)", "Worker", "leidy.garcia", "leidy123"),
        ("Yusney Garcia (Profesional TH)", "Worker", "yusney.garcia", "yusney123"),
        ("Yenny Cardenas (Profesional TH)", "Worker", "yenny.cardenas", "yenny123"),
    ]

    for row_idx, data in enumerate(users_data, 5):
        ws_cred.row_dimensions[row_idx].height = 22
        for col_idx, val in enumerate(data, 1):
            cell = ws_cred.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = font_regular
            cell.border = thin_border
            # Styles per column
            if col_idx == 2:
                cell.alignment = align_center
                if val == "Admin":
                    cell.font = font_bold
            elif col_idx in [3, 4]:
                cell.alignment = align_center
                cell.fill = zebra_fill
            else:
                cell.alignment = align_left

    # ─────────────────────────────────────────────────────────────────────────
    # HOJA 2: CHECKLIST
    # ─────────────────────────────────────────────────────────────────────────
    ws_check = wb.create_sheet(title="2. Checklist de Pruebas")
    ws_check.views.sheetView[0].showGridLines = True

    # Title block
    ws_check["A1"] = "📋 Lista de Chequeo y Pruebas del Sistema"
    ws_check["A1"].font = font_title
    ws_check.row_dimensions[1].height = 30

    ws_check["A2"] = "Sigue estos pasos ordenadamente perfil por perfil para verificar la integridad del sistema:"
    ws_check["A2"].font = font_italic
    ws_check.row_dimensions[2].height = 20

    # Table Headers
    headers_check = ["Sección / Perfil", "Paso", "Descripción del Paso", "Verificación Esperada (Criterio de Aceptación)", "Estado (Pendiente / OK)", "Notas de Prueba"]
    for col_idx, h in enumerate(headers_check, 1):
        cell = ws_check.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = green_header_fill
        cell.alignment = align_center
        cell.border = thick_bottom_border
    ws_check.row_dimensions[4].height = 28

    # Checklist items
    checklist_data = [
        # Sección 1
        ("1. Acceso y Sidebar por Rol", "Paso 1.1", 
         "Inicia sesión como Administrador (admin / admin123).", 
         "El menú lateral del sidebar debe mostrar exactamente 5 módulos: Inicio, Estructura TH, Exportar Reporte, Seguimiento de Tareas, y Configuración."),
        ("1. Acceso y Sidebar por Rol", "Paso 1.2", 
         "Inicia sesión como Supervisor (lilia.nocua / lilia123).", 
         "El menú lateral debe mostrar solo 2 módulos: Inicio (con panel de equipo) y Gestión de Tareas (para supervisión de tareas)."),
        ("1. Acceso y Sidebar por Rol", "Paso 1.3", 
         "Inicia sesión como Trabajador (laura.cely o claudia.garcia / contraseña cely123 o claudia123).", 
         "El menú debe contener única y exclusivamente el módulo 'Mis Tareas'."),
         
        # Sección 2
        ("2. Flujo del Trabajador (Worker)", "Paso 2.1", 
         "Inicia sesión como Trabajador (ej. rosaura.ramirez).", 
         "Las tareas asignadas deben mostrarse ordenadas por prioridad (Vencidas, Esta semana, Restantes). Las tareas ya completadas deben listarse al final en sección colapsada."),
        ("2. Flujo del Trabajador (Worker)", "Paso 2.2", 
         "Selecciona una tarea en proceso y mueve el slider a un valor intermedio (ej. 70%). Escribe una observación y agrega un enlace en el campo de evidencias. Guarda.", 
         "El progreso de la tarea se actualiza y el estado de la tarea cambia automáticamente a 'En Proceso' con color de semáforo correspondiente."),
        ("2. Flujo del Trabajador (Worker)", "Paso 2.3", 
         "Modifica la misma tarea al 100% de progreso y dale guardar.", 
         "El estado cambia automáticamente a 'Cumplida', los campos se bloquean (solo lectura) y la tarea se traslada al historial colapsado de tareas cumplidas."),

        # Sección 3
        ("3. Flujo del Supervisor", "Paso 3.1", 
         "Inicia sesión como Supervisor (lilia.nocua). Revisa la pestaña 'Mi Equipo'.", 
         "Se visualiza la tabla consolidada de avance de todas las funcionarias (Rosaura, Laura, Claudia, etc.), con totales de tareas cumplidas y vencidas, progreso medio y semáforo por funcionaria."),
        ("3. Flujo del Supervisor", "Paso 3.2", 
         "Pestaña 'Gestión de Tareas': Usa los filtros horizontales por responsable y por política.", 
         "La grilla de tareas se filtra instantáneamente en tiempo real según el criterio seleccionado."),
        ("3. Flujo del Supervisor", "Paso 3.3", 
         "Haz clic en editar (ícono del lápiz ✏️) en una tarea que esté Vencida. Intenta guardar el formulario dejando vacía la 'Justificación de Retraso'.", 
         "El sistema debe rechazar el guardado y mostrar una alerta roja indicando que la justificación de retraso es obligatoria por estar fuera de plazo."),
        ("3. Flujo del Supervisor", "Paso 3.4", 
         "Escribe la justificación de retraso y haz clic en Guardar.", 
         "El cambio se guarda con éxito, se registra en el log y el modal de edición se cierra correctamente."),

        # Sección 4
        ("4. Indicadores y Visualización", "Paso 4.1", 
         "Inicia sesión como Admin o Supervisor y ve al Dashboard de Inicio.", 
         "El velocímetro principal de avance y los gráficos de barra de Plotly se muestran de forma responsiva con los porcentajes reales."),
        ("4. Indicadores y Visualización", "Paso 4.2", 
         "Revisa la sección 'Desglose Estratégico' (gráfico Treemap).", 
         "Los títulos de las actividades/estrategias dentro del gráfico son visibles y legibles (se truncan automáticamente a un máximo de 50 caracteres para evitar fuentes microscópicas)."),
        ("4. Indicadores y Visualización", "Paso 4.3", 
         "Pasa el cursor (hover) por encima de los bloques del Treemap.", 
         "Se despliega un cuadro informativo mostrando la estrategia correspondiente y su avance porcentual sin errores."),

        # Sección 5
        ("5. Auditoría y Bloqueo", "Paso 5.1", 
         "Inicia sesión como admin y ve a Configuración -> pestaña Continuidad.", 
         "Se visualiza la tabla 'Historial de Cambios' con todos los registros exactos de las pruebas anteriores (usuario, campo modificado, valor antiguo y nuevo)."),
        ("5. Auditoría y Bloqueo", "Paso 5.2", 
         "Haz clic en el botón '🔒 Cerrar y Bloquear Periodo' para la vigencia activa (2025).", 
         "El estado de la vigencia se muestra como Bloqueado."),
        ("5. Auditoría y Bloqueo", "Paso 5.3", 
         "Entra como Worker (rosaura.ramirez) e intenta editar una tarea o cargar evidencia.", 
         "Todos los botones de guardado y campos se muestran inhabilitados en modo Solo Lectura, impidiendo modificaciones por cierre de auditoría."),
        ("5. Auditoría y Bloqueo", "Paso 5.4", 
         "Regresa como admin, ve a Continuidad y haz clic en '🔓 Desbloquear Periodo'.", 
         "El periodo se desbloquea exitosamente y se reestablece la capacidad de escritura."),

        # Sección 6
        ("6. Exportación de Reportes", "Paso 6.1", 
         "Ve al módulo 'Exportar Reporte', define el año 2025 y tipo 'Año Completo'.", 
         "Se muestra el rango del 01/01/2025 al 31/12/2025."),
        ("6. Exportación de Reportes", "Paso 6.2", 
         "Ve a la pestaña 'Informe Word' y descarga el documento .docx.", 
         "Se genera el documento Word corporativo con portada en verde, introducción de gestión, firmas y tablas de compromisos estilizadas cebra con hipervínculos de evidencias en verde."),
        ("6. Exportación de Reportes", "Paso 6.3", 
         "Ve a la pestaña 'Excel Ejecutivo' y descarga el libro .xlsx.", 
         "El archivo abre con 4 hojas: Resumen por política, Detalle jerárquico completo, Avance por equipo y Trazabilidad de cambios, con formato corporativo."),
        ("6. Exportación de Reportes", "Paso 6.4", 
         "Ve a la pestaña 'PDF Institucional' y descarga el documento .pdf.", 
         "El PDF se genera con el logotipo oficial de Unitrópico en el membrete superior, formato en escala de colores y firmas al final."),
    ]

    for row_idx, data in enumerate(checklist_data, 5):
        # Determine row height based on text lengths
        max_len = max(len(data[2]), len(data[3]))
        ws_check.row_dimensions[row_idx].height = 40 if max_len > 80 else 24

        # Alternate section backgrounds to group visually
        sec_name = data[0]
        row_fill = zebra_fill if int(sec_name[0]) % 2 == 0 else white_fill

        for col_idx, val in enumerate(data, 1):
            cell = ws_check.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border
            cell.fill = row_fill
            
            # Alignments and fonts
            if col_idx in [1, 2]:
                cell.font = font_bold
                cell.alignment = align_center
            elif col_idx in [3, 4]:
                cell.font = font_regular
                cell.alignment = align_left
            
        # Empty cells for Status and Notes (User input)
        status_cell = ws_check.cell(row=row_idx, column=5)
        status_cell.font = font_bold
        status_cell.border = thin_border
        status_cell.fill = light_yellow_fill
        status_cell.alignment = align_center

        notes_cell = ws_check.cell(row=row_idx, column=6)
        notes_cell.font = font_regular
        notes_cell.border = thin_border
        notes_cell.fill = light_yellow_fill
        notes_cell.alignment = align_left

    # ─────────────────────────────────────────────────────────────────────────
    # Auto-fit columns for all sheets
    # ─────────────────────────────────────────────────────────────────────────
    for sheet in wb.worksheets:
        for col in sheet.columns:
            # We skip auto-sizing for the checklist text columns to set custom wraps instead
            if sheet.title == "2. Checklist de Pruebas":
                sheet.column_dimensions['A'].width = 25
                sheet.column_dimensions['B'].width = 12
                sheet.column_dimensions['C'].width = 45
                sheet.column_dimensions['D'].width = 50
                sheet.column_dimensions['E'].width = 25
                sheet.column_dimensions['F'].width = 40
                break
                
            max_len = 0
            for cell in col:
                # Avoid checking formula values or titles at row 1
                if cell.row in [1, 2]:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to project root
    output_path = os.path.join(ROOT_DIR, "PRUEBAS_CHECKLIST.xlsx")
    wb.save(output_path)
    print(f"Éxito! Checklist Excel creado en: {output_path}")

if __name__ == "__main__":
    build_excel_checklist()
