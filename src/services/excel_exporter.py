"""
excel_exporter.py — Fase 2.5
Genera un reporte Excel ejecutivo con 3 hojas:
  1. Resumen     — Métricas consolidadas por Política
  2. Detalle     — Todas las tareas con estado, avance y evidencias
  3. Equipo      — Avance por funcionaria
  4. Historial   — Últimos 200 cambios (si existen)

Uso:
    from src.services.excel_exporter import build_excel_report
    bytes_io = build_excel_report(db)
    st.download_button("Descargar Excel", bytes_io, "reporte_th.xlsx", ...)
"""

import io
from datetime import datetime
from sqlalchemy.orm import Session, joinedload
from src.models.entities import PlanMacro, Policy, StrategicItem, Activity, Task, Responsible, ChangeLog

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ─────────────────────────────────────────────────────────────────────────────
# Estilos
# ─────────────────────────────────────────────────────────────────────────────
GREEN    = "FF00594E"   # Verde institucional
GOLD     = "FFB5A160"   # Dorado institucional
WHITE    = "FFFFFFFF"
RED      = "FFEF4444"
AMBER    = "FFF59E0B"
EMERALD  = "FF10B981"
LIGHT_GRAY = "FFF8FAFC"
HEADER_GRAY = "FFE2E8F0"

def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, color="FF1E293B", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def _border():
    s = Side(style="thin", color="FFE2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, row_num, headers, col_widths=None):
    """Write a styled header row."""
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=col, value=h)
        c.fill      = _fill(GREEN)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.alignment = _center()
        c.border    = _border()
        if col_widths: ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]

def _semaforo_fill(progress):
    if progress >= 80:  return _fill(EMERALD)
    if progress >= 50:  return _fill(AMBER)
    return _fill(RED)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 1: Resumen por Política
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_resumen(wb, macro: PlanMacro):
    ws = wb.create_sheet("1. Resumen")
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"] = f"INFORME DE GESTIÓN — {macro.name} ({macro.year})"
    ws["A1"].fill      = _fill(GREEN)
    ws["A1"].font      = _font(bold=True, color=WHITE, size=14)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} — Avance consolidado: {macro.progress:.1f}%"
    ws["A2"].fill      = _fill(GOLD)
    ws["A2"].font      = _font(bold=True, color=WHITE, size=11)
    ws["A2"].alignment = _center()

    # Encabezados
    headers = ["Política", "Estrategias", "Actividades", "Tareas Total",
               "✅ Cumplidas", "🔴 Vencidas", "Avance %"]
    widths  = [42, 14, 14, 14, 14, 14, 14]
    _header_row(ws, 4, headers, widths)

    now = datetime.now()
    totals = {"est": 0, "act": 0, "tar": 0, "cum": 0, "ven": 0}
    for r, pol in enumerate(macro.policies, 5):
        tasks_all = [t for si in pol.strategic_items for act in si.activities for t in act.tasks]
        cumplidas = sum(1 for t in tasks_all if t.status == "Cumplida")
        vencidas  = sum(1 for t in tasks_all if t.status != "Cumplida" and t.end_date and t.end_date < now)
        acts_all  = [act for si in pol.strategic_items for act in si.activities]

        row_data = [pol.name, len(pol.strategic_items), len(acts_all),
                    len(tasks_all), cumplidas, vencidas, round(pol.progress, 1)]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill      = _fill(LIGHT_GRAY if r % 2 == 0 else WHITE)
            c.font      = _font(size=10)
            c.alignment = _left() if col == 1 else _center()
            c.border    = _border()

        # Colorear avance
        pct_cell = ws.cell(row=r, column=7)
        pct_cell.fill = _semaforo_fill(pol.progress)
        pct_cell.font = _font(bold=True, color=WHITE, size=10)

        totals["est"] += len(pol.strategic_items)
        totals["act"] += len(acts_all)
        totals["tar"] += len(tasks_all)
        totals["cum"] += cumplidas
        totals["ven"] += vencidas

    # Fila de totales
    tf = len(macro.policies) + 5
    totals_data = ["TOTAL", totals["est"], totals["act"], totals["tar"],
                   totals["cum"], totals["ven"], round(macro.progress, 1)]
    for col, val in enumerate(totals_data, 1):
        c = ws.cell(row=tf, column=col, value=val)
        c.fill      = _fill(GOLD)
        c.font      = _font(bold=True, color=WHITE, size=11)
        c.alignment = _center()
        c.border    = _border()


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 2: Detalle de Tareas
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_detalle(wb, macro: PlanMacro):
    ws = wb.create_sheet("2. Detalle Tareas")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    ws["A1"] = f"DETALLE DE TAREAS — {macro.name} ({macro.year})"
    ws["A1"].fill = _fill(GREEN)
    ws["A1"].font = _font(bold=True, color=WHITE, size=13)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["Política", "Estrategia", "Actividad", "Tarea",
               "Responsable", "Estado", "Avance %",
               "Inicio", "Fin", "Observaciones", "Evidencias"]
    widths  = [28, 28, 30, 38, 22, 14, 10, 12, 12, 35, 45]
    _header_row(ws, 3, headers, widths)

    now  = datetime.now()
    row  = 4
    for pol in macro.policies:
        for si in pol.strategic_items:
            for act in si.activities:
                for t in act.tasks:
                    resp = ", ".join(r.name for r in t.responsibles) or "Sin asignar"
                    evs  = "; ".join(ev.url for ev in t.evidences) if t.evidences else "—"
                    row_data = [
                        pol.name, si.name, act.name, t.name, resp, t.status,
                        round(t.progress, 1),
                        t.start_date.strftime("%d/%m/%Y") if t.start_date else "—",
                        t.end_date.strftime("%d/%m/%Y")   if t.end_date   else "—",
                        t.observations or "—",
                        evs
                    ]
                    is_vencida = (t.status != "Cumplida" and t.end_date and t.end_date < now)
                    bg = _fill("FFFEF2F2") if is_vencida else _fill(LIGHT_GRAY if row % 2 == 0 else WHITE)

                    for col, val in enumerate(row_data, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.fill      = bg
                        c.font      = _font(size=9)
                        c.alignment = _left() if col in (1, 2, 3, 4, 10, 11) else _center()
                        c.border    = _border()

                    # Avance con color
                    pct_cell = ws.cell(row=row, column=7)
                    pct_cell.fill = _semaforo_fill(t.progress)
                    pct_cell.font = _font(bold=True, color=WHITE, size=9)

                    row += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{row - 1}"


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 3: Avance por Equipo
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_equipo(wb, macro: PlanMacro, db: Session):
    ws = wb.create_sheet("3. Equipo")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = f"AVANCE POR FUNCIONARIA — {macro.name} ({macro.year})"
    ws["A1"].fill = _fill(GREEN)
    ws["A1"].font = _font(bold=True, color=WHITE, size=13)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["Funcionaria", "Cargo", "Total Tareas", "✅ Cumplidas",
               "⚙️ En Proceso", "🔴 Pendientes", "🚫 Vencidas", "Avance %"]
    widths  = [32, 30, 14, 14, 14, 14, 14, 14]
    _header_row(ws, 3, headers, widths)

    now = datetime.now()
    responsibles = db.query(Responsible).options(joinedload(Responsible.tasks)).all()
    row = 4
    for r in responsibles:
        tasks_y = [
            t for t in r.tasks
            if t.activity and t.activity.strategic_item
            and t.activity.strategic_item.policy
            and t.activity.strategic_item.policy.plan_macro_id == macro.id
        ]
        if not tasks_y:
            continue
        total    = len(tasks_y)
        cumpl    = sum(1 for t in tasks_y if t.status == "Cumplida")
        en_proc  = sum(1 for t in tasks_y if t.status == "En Proceso")
        pend     = sum(1 for t in tasks_y if t.status == "Pendiente")
        vencidas = sum(1 for t in tasks_y if t.status != "Cumplida" and t.end_date and t.end_date < now)
        avg      = sum(t.progress for t in tasks_y) / total

        row_data = [r.name, r.role, total, cumpl, en_proc, pend, vencidas, round(avg, 1)]
        bg = _fill(LIGHT_GRAY if row % 2 == 0 else WHITE)
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = bg
            c.font      = _font(size=10)
            c.alignment = _left() if col <= 2 else _center()
            c.border    = _border()

        pct_c = ws.cell(row=row, column=8)
        pct_c.fill = _semaforo_fill(avg)
        pct_c.font = _font(bold=True, color=WHITE, size=10)
        row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 4: Historial de cambios
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_historial(wb, db: Session):
    logs = db.query(ChangeLog).order_by(ChangeLog.changed_at.desc()).limit(500).all()
    if not logs:
        return

    ws = wb.create_sheet("4. Historial")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "HISTORIAL DE CAMBIOS"
    ws["A1"].fill = _fill(GREEN)
    ws["A1"].font = _font(bold=True, color=WHITE, size=13)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["Fecha y Hora", "Usuario", "Tarea", "Campo", "Antes", "Después"]
    widths  = [18, 18, 45, 12, 22, 22]
    _header_row(ws, 3, headers, widths)

    for row, l in enumerate(logs, 4):
        row_data = [
            l.changed_at.strftime("%d/%m/%y %H:%M"),
            l.changed_by,
            l.entity_name,
            l.field_changed,
            l.old_value or "—",
            l.new_value or "—",
        ]
        bg = _fill(LIGHT_GRAY if row % 2 == 0 else WHITE)
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = bg
            c.font      = _font(size=9)
            c.alignment = _left() if col in (2, 3) else _center()
            c.border    = _border()


# ─────────────────────────────────────────────────────────────────────────────
# Función principal pública
# ─────────────────────────────────────────────────────────────────────────────
def build_excel_report(db: Session, year: int = None) -> io.BytesIO:
    """
    Generates a complete Excel report as BytesIO.
    Returns None if openpyxl is not installed.
    """
    if not OPENPYXL_OK:
        return None

    # Cargar el plan macro del año solicitado (o el más reciente)
    query = db.query(PlanMacro).options(
        joinedload(PlanMacro.policies)
        .joinedload(Policy.strategic_items)
        .joinedload(StrategicItem.activities)
        .joinedload(Activity.tasks)
        .joinedload(Task.responsibles),
        joinedload(PlanMacro.policies)
        .joinedload(Policy.strategic_items)
        .joinedload(StrategicItem.activities)
        .joinedload(Activity.tasks)
        .joinedload(Task.evidences)
    ).order_by(PlanMacro.year.desc())

    if year:
        query = query.filter(PlanMacro.year == year)
    macro = query.first()

    if not macro:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    _sheet_resumen(wb, macro)
    _sheet_detalle(wb, macro)
    _sheet_equipo(wb, macro, db)
    _sheet_historial(wb, db)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
